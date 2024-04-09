import tkinter as tk
from tkinter import ttk, messagebox
import customtkinter as ctk
from crypto_utils import derive_key, encrypt_password, decrypt_password
from excel_utils import add_password_to_excel
from openpyxl import load_workbook


class LoginWindow:
    def __init__(self, parent, login_function):
        self.parent = parent
        self.login_function = login_function
        self.setup_window()

    def setup_window(self):
        self.top = ctk.CTkToplevel(self.parent)
        self.top.geometry("450x300")
        self.top.title("Login Page")
        self.top.resizable(False, False)

        login_frame_width = self.top.winfo_width()
        login_frame_height = self.top.winfo_height()
        login_frame = ctk.CTkFrame(self.top, width=login_frame_width, height=login_frame_height, corner_radius=10)
        login_frame.place(x=125, y=50)

        welcome_label = ctk.CTkLabel(login_frame, text="Welcome Back!")
        welcome_label.pack(pady=12)

        password_entry = ctk.CTkEntry(login_frame, placeholder_text="Password:", show="*")
        password_entry.pack(pady=10, padx=20, fill='x')

        login_button = ctk.CTkButton(login_frame, text="Login", command=lambda: self.on_login(password_entry.get()))
        login_button.pack(pady=20)

    def on_login(self, master_password):
        login_success = self.login_function(master_password)
        if login_success:
            self.top.destroy()
        else:
            messagebox.showerror("Login Failed", "The provided master password is incorrect.")


class PasswordManagerGUI:
    def __init__(self, root, master_password):
        self.root = root
        self.master_password = master_password
        self.key = derive_key(master_password)
        ctk.set_appearance_mode("light")
        self.root.title("Password Manager")
        self.style = ttk.Style()
        self.create_widgets()

    def create_widgets(self):
        labels_texts = ["Site Name:", "Username:", "Password:", "Notes:"]
        self.entries = []
        self.vars = {text: tk.StringVar() for text in labels_texts}

        for i, text in enumerate(labels_texts):
            label = ttk.Label(self.root, text=text)
            label.grid(column=0, row=i, padx=10, pady=10, sticky="W")

            entry = ttk.Entry(self.root, textvariable=self.vars[text])
            self.vars[text].trace("w", lambda name, index, mode, sv=self.vars[text], field=text: self.limit_input(sv, field))
            entry.grid(column=1, row=i, padx=10, pady=10, sticky="EW")
            self.entries.append(entry)

        add_button = ctk.CTkButton(self.root, text="Add Password", command=self.add_password)
        add_button.grid(column=0, row=4, columnspan=2, padx=10, pady=20)

        view_button = ctk.CTkButton(self.root, text="View Passwords", command=self.view_passwords)
        view_button.grid(column=0, row=5, columnspan=2, padx=10, pady=15)

        self.root.grid_columnconfigure(1, weight=1)

    def limit_input(self, sv, field):
        content = sv.get()
        if len(content) > 30:
            sv.set(content[:30])
            tk.messagebox.showwarning("Warning", f"{field} cannot exceed 30 characters.")

    def add_password(self):
        site, username, password, notes = (entry.get() for entry in self.entries)
        encrypted_password, salt = encrypt_password(password,
                                                    self.master_password)
        add_password_to_excel(site, username, encrypted_password, salt, notes)

        for entry in self.entries:
            entry.delete(0, tk.END)
        print(f"Password for {site} added successfully.")

    def view_passwords(self):
        confirm = messagebox.askokcancel("Warning","All passwords will be displayed in plain text. Do you want to proceed?")
        if not confirm:
            return  # User canceled, do not proceed
        filename = "passwords.xlsx"
        wb = load_workbook(filename)
        sheet = wb.active

        new_window = tk.Toplevel(self.root)
        new_window.title("View Stored Passwords")
        new_window.geometry("700x500")
        new_window.resizable(False, False)

        canvas = tk.Canvas(new_window)
        v_scrollbar = tk.Scrollbar(new_window, orient="vertical", command=canvas.yview)
        h_scrollbar = tk.Scrollbar(new_window, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        scrollable_frame = tk.Frame(canvas)
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        canvas.pack(side="left", fill="both", expand=True)

        sort_order = ttk.Combobox(scrollable_frame, values=["Ascending", "Descending"], state="readonly")
        sort_order.grid(row=0, column=0, padx=10, pady=5)
        sort_order.set("Ascending")

        sort_button = ctk.CTkButton(scrollable_frame, text="Sort",
                                    command=lambda: self.fill_passwords(sheet, scrollable_frame, "Platform",
                                                                        sort_order.get().lower()))
        sort_button.grid(row=0, column=1, padx=10, pady=5)

        self.fill_passwords(sheet, scrollable_frame)

    def fill_passwords(self, sheet, window, sort_by=None, sort_order='ascending'):
        for widget in window.grid_slaves():
            if int(widget.grid_info()["row"]) > 0:
                widget.grid_forget()

        passwords = [
            (row[0], row[1], decrypt_password(row[2], self.master_password, row[3]), row[4] or "")
            for row in sheet.iter_rows(min_row=1, values_only=True)
        ]

        if sort_by:
            sort_index = ["Platform", "Username", "Notes"].index(sort_by)
            passwords.sort(key=lambda x: x[sort_index], reverse=(sort_order == 'descending'))

        for i, (site, username, decrypted_password, notes) in enumerate(passwords, start=1):
            tk.Label(window, text=site, fg='black').grid(row=i, column=0, padx=10, pady=5, sticky='w')
            tk.Label(window, text=username, fg='black').grid(row=i, column=1, padx=10, pady=5, sticky='w')
            tk.Label(window, text=decrypted_password, fg='black').grid(row=i, column=2, padx=10, pady=5, sticky='w')
            tk.Label(window, text=notes, fg='black').grid(row=i, column=3, padx=10, pady=5, sticky='w')

            delete_button = tk.Button(window, text="X", command=lambda rn=i: self.confirm_delete(sheet, window, rn))
            delete_button.grid(row=i, column=4, padx=10, pady=5, sticky='w')

    def confirm_delete(self, sheet, window, excel_row):
        confirm = tk.messagebox.askyesno("Confirm Deleting", "This will delete the credentials, proceed?")
        if confirm:
            sheet.delete_rows(excel_row)
            sheet.parent.save("passwords.xlsx")
            self.fill_passwords(sheet, window)

    def apply_sort(self, sheet, window, sort_by, sort_order):
        self.fill_passwords(sheet, window, sort_by=sort_by, sort_order=sort_order)