from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def add_password_to_excel(site, username, encrypted_password, salt, notes, filename="passwords.xlsx"):
    """Add a new password entry to the Excel file, including the salt used for encryption."""
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
    else:
        sheet = wb.active

    # Save encrypted password and hex-encoded salt
    sheet.append([site, username, encrypted_password, salt, notes])
    wb.save(filename)